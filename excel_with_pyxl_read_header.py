'''
This is the second version which can read an excel and create a fastload script.
DATA File is created without hardcoding each column
List of fields are read from the header row of excel
Columns names for fastload are used from header
'''


def create_data_file(sheet,outfile):
    maxcol = sheet.max_column + 1
    for r in range (1, maxcol):
        col = sheet.cell (row=1, column=r)
        header.append (col.value.replace (" ", "_"))
    print (header)

    rec_num = 1

    for row_cells in sheet.iter_rows(min_row=2):
        col_itr = 1
        outfile.write (str(rec_num)+"~")
        for cell in row_cells:
            if cell.value is None:
                cell.value = "#"
            write_value = str (" ".join (str (cell.value).split ()))
            if col_itr < int (sheet.max_column):
                print (write_value, end="~")
                text_to_file = str (write_value) + "~"
                outfile.write (text_to_file)
            else:
                print (write_value, end="\n")
                text_to_file = str (write_value) + "\n"
                outfile.write (text_to_file)

            col_itr = col_itr + 1
            rec_num=rec_num+1


def create_fl_script():
    # CREATING A FASTLOAD SCRIPT
    print ("Creating FASTLOAD .. ")


    fl_file.write (
        "SESSIONS 5 ;" + "\n" + "TENACITY 5;" + "\n" + ".<HOSTNAME>/<USER>,<PWD>" + "\n")
    fl_file.write ("DROP TABLE DBNAME.stg_table;" + "\n")
    fl_file.write ("DROP TABLE DBNAME.stg_table_err1;" + "\n")
    fl_file.write ("DROP TABLE DBNAME.stg_table_err2;" + "\n" + "\n")
    fl_file.write ("CREATE TABLE DBNAME.stg_table " + "\n"
                   "(" + "\n" +
                   "rec_num INTEGER," + "\n"
                   "country varchar(20)," + "\n"
                   "ee_case_id varchar(200)," + "\n"
                   "case_closure_date varchar(20)," + "\n"
                   "source varchar(100) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "product varchar(100)," + "\n"
                    "rcacat varchar(100) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "rcatrend1 varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "rcatrend2 varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "rcatrend3 varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "complaint_dtl varchar(8000) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "cust_id varchar(100) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "cust_name varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC" + "\n"
                   + ")" + "Primary index(country,ee_case_id,source,product);" + "\n"
                   )
    fl_file.write ("\n")
    fl_file.write ("DEFINE " + "\n"
                   "rec_num (integer)," + "\n"
                   "country (varchar(20))," + "\n"
                   "ee_case_id (varchar(200))," + "\n"
                   "case_closure_date (varchar(20))," + "\n"
                   "source (varchar(100))," + "\n"
                   "product (varchar(100))," + "\n"
                   "rcacat (varchar(100))," + "\n"
                   "rcatrend1 (varchar(300))," + "\n"
                   "rcatrend2 (varchar(300))," + "\n"
                   "rcatrend3 (varchar(300))," + "\n"
                   "complaint_dtl (varchar(8000))," + "\n"
                   "cust_id (varchar(100))," + "\n"
                   "cust_name (varchar(300))" + "\n"
                   )
    fl_file.write ("\n")
    fl_file.write ("FILE=" + "\"" + "A:\Ash Learning\Python Udemy courseera\Ash_Exercises\data_load.txt" + "\"" + ";\n")
    fl_file.write ("\n")
    fl_file.write ("BEGIN LOADING  DBNAME.stg_table , DBNAME.stg_table_err1," \
                   "DBNAME.stg_table_err2;" + "\n")
    fl_file.write ("\n")

    fl_file.write ("INSERT INTO  DBNAME.stg_table(" + "\n" 
                   "REC_NUM," + "\n" 
                   "COUNTRY," + "\n" 
                   "EE_CASE_ID," + "\n" 
                   "CASE_CLOSURE_DATE," + "\n" 
                   "SOURCE," + "\n" 
                   "PRODUCT," + "\n"
                   "RCACAT," + "\n"
                   "RCATREND1," + "\n"
                   "RCATREND2," + "\n"
                   "RCATREND3," + "\n"
                   "COMPLAINT_DTL," + "\n"
                   "CUST_ID," + "\n"
                   "CUST_NAME" + "\n)"
                   )
    fl_file.write ("Values (" + "\n" 
                   ":rec_num," + "\n" 
                   ":region," + "\n" 
                   ":country," + "\n" 
                   ":ee_case_id," + "\n" 
                   ":case_closure_date," + "\n" 
                   ":source," + "\n" 
                   ":product," + "\n" 
                   ":rcacat," + "\n" 
                   ":rcatrend1," + "\n" 
                   ":rcatrend2," + "\n" 
                   ":rcatrend3," + "\n" 
                   ":complaint_dtl," + "\n" 
                   ":cust_id," + "\n" 
                   ":cust_name);" + "\n"
                   )

    fl_file.write ("END LOADING;" + "\n")
    fl_file.write ("\n")
    fl_file.write ("LOG OFF;" + "\n")
    print ("FASTLOAD SCRIPT COMPLETE. ")


'''
for key , value in rca_data.items():
    print ("Key is " + str(key) + " --> Value is " + str(value).lstrip('[').rstrip(']'))

'''


if __name__ == "__main__":
    import openpyxl
    import sys

    print ("hello")
    wb = openpyxl.load_workbook ('A:\Ash Learning\Python Udemy courseera\Ash_Exercises\Openpyxl-Excel to Txt\pyxl_input.xlsx')
    in_sheet = wb['Jan_Data']
    rca_data = {}
    i = 1
    arg_outfile = open ('A:\Ash Learning\Python Udemy courseera\Ash_Exercises\Openpyxl-Excel to Txt\data_load.txt', 'w', encoding="utf-8")
    fl_file = open ('data_load.fld', 'w', encoding="utf-8")
    print ("Max col is " + str (in_sheet.max_column))
    header=[]

    create_data_file(in_sheet,arg_outfile)
    #create_fl_script()


