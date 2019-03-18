'''
This is the first version which can read an excel and create a fastload script.
List of fields from excel are hardcoded
Columns names for fastload are hardcoded
'''

import openpyxl
import sys
print ("hello")
wb = openpyxl.load_workbook('pyxl_input.xlsx')
sheet= wb['Jan_Data']
rca_data= {}
i=1
outfile = open('data_load.txt','w',encoding="utf-8")
fl_file = open ('data_load.fld', 'w', encoding="utf-8")
print("Max col is " + str(sheet.max_column))

def create_data_file(x):
    for row in range(2,sheet.max_row+1):
        region=sheet['A' + str(row)].value
        country=sheet['B' + str(row)].value
        ee_case_id = sheet['C' + str (row)].value
        case_closure_date=str(sheet['D' + str (row)].value).replace("\n",'')
        source=sheet['E' + str (row)].value
        product = sheet['F' + str (row)].value
        rcacat = sheet['G' + str (row)].value
        rcatrend1 = sheet['H' + str (row)].value
        rcatrend2= sheet['I' + str (row)].value
        rcatrend3 = " ".join(str(sheet['J' + str (row)].value).split())
        complaint_dtl = " ".join (str (sheet['K' + str (row)].value).split() )
        cust_id = " ".join (str (sheet['L' + str (row)].value).split ())
        cust_name = " ".join (str (sheet['M' + str (row)].value).split ())
        #ixn_ts = str(sheet['B' + str (row)].value).replace("\n",'')
        #ixn_ts=" " .join(str(sheet['B' + str (row)].value))
        #comp_dtl = sheet['C' + str (row)].value
        #comp_dtl = " ".join (str (sheet['C' + str (row)].value).split() )
        print (str(x) + region + " " + str(country)+ " " + str(ee_case_id) +" " + str(case_closure_date) + " " + str(source) )
        #rca_data[i] = [str(ixn_id)+'~',str(ixn_ts)+'~',str(comp_dtl)+'~']
        outfile.write(str(x)+"~" +str(region) + "~" + str(country)+ "~" + str(ee_case_id) +"~" + str(case_closure_date) + "~" + str(source)+
                      str(product) + "~" + str (rcacat) + "~" + str (rcatrend1) + "~" + str (rcatrend2) + "~" + str (rcatrend3)+"~"+
                      str(complaint_dtl)+"~"+str(cust_id)+"~" + str(cust_name)+'\n')
        x=x+1


def create_fl_script():
    # CREATING A FASTLOAD SCRIPT
    print ("Creating FASTLOAD .. ")


    fl_file.write (
        "SESSIONS 5 ;" + "\n" + "TENACITY 5;" + "\n" + ".logon <DATABASE>/<USERNMAE>,<PWD>" + "\n")
    fl_file.write ("DROP TABLE DBNAME.stg_table;" + "\n")
    fl_file.write ("DROP TABLE DBNAME.stg_table_err1;" + "\n")
    fl_file.write ("DROP TABLE DBNAME.stg_table_err2;" + "\n" + "\n")
    fl_file.write ("CREATE TABLE DBNAME.stg_table " + "\n"
                   "(" + "\n" +
                   "rec_num INTEGER," + "\n"
                   "country varchar(20)," + "\n"
                   "ee_case_id varchar(200)," + "\n"
                   "case_closure_date varchar(20)," + "\n"
                   "source varchar(100) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "product varchar(100)," + "\n"
                    "rcacat varchar(100) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "rcatrend1 varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "rcatrend2 varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "rcatrend3 varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "complaint_dtl varchar(8000) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "cust_id varchar(100) CHARACTER SET UNICODE NOT CASESPECIFIC," + "\n"
                    "cust_name varchar(300) CHARACTER SET UNICODE NOT CASESPECIFIC" + "\n"
                   + ")" + "Primary index(country,ee_case_id,source,product);" + "\n"
                   )
    fl_file.write ("\n")
    fl_file.write ("SET RECORD VARTEXT = ~;")
    fl_file.write ("\n")
    fl_file.write ("DEFINE " + "\n"
                   "rec_num (integer)," + "\n"
                   "country (varchar(20))," + "\n"
                   "ee_case_id (varchar(200))," + "\n"
                   "case_closure_date (varchar(20))," + "\n"
                   "source (varchar(100))," + "\n"
                   "product (varchar(100))," + "\n"
                   "rcacat (varchar(100))," + "\n"
                   "rcatrend1 (varchar(300))," + "\n"
                   "rcatrend2 (varchar(300))," + "\n"
                   "rcatrend3 (varchar(300))," + "\n"
                   "complaint_dtl (varchar(8000))," + "\n"
                   "cust_id (varchar(100))," + "\n"
                   "cust_name (varchar(300))" + "\n"
                   )
    fl_file.write ("\n")
    fl_file.write ("FILE=" + "\"" + "A:\Ash Learning\Python Udemy courseera\Ash_Exercises\data_load.txt" + "\"" + ";\n")
    fl_file.write ("\n")
    fl_file.write ("BEGIN LOADING  DBNAME.stg_table , DBNAME.stg_table_err1," \
                   "DBNAME.stg_table_err2;" + "\n")
    fl_file.write ("\n")

    fl_file.write ("INSERT INTO  DBNAME.stg_table(" + "\n" 
                   "REC_NUM," + "\n" 
                   "COUNTRY," + "\n" 
                   "EE_CASE_ID," + "\n" 
                   "CASE_CLOSURE_DATE," + "\n" 
                   "SOURCE," + "\n" 
                   "PRODUCT," + "\n"
                   "RCACAT," + "\n"
                   "RCATREND1," + "\n"
                   "RCATREND2," + "\n"
                   "RCATREND3," + "\n"
                   "COMPLAINT_DTL," + "\n"
                   "CUST_ID," + "\n"
                   "CUST_NAME" + "\n)"
                   )
    fl_file.write ("Values (" + "\n" 
                   ":rec_num," + "\n" 
                   ":region," + "\n" 
                   ":country," + "\n" 
                   ":ee_case_id," + "\n" 
                   ":case_closure_date," + "\n" 
                   ":source," + "\n" 
                   ":product," + "\n" 
                   ":rcacat," + "\n" 
                   ":rcatrend1," + "\n" 
                   ":rcatrend2," + "\n" 
                   ":rcatrend3," + "\n" 
                   ":complaint_dtl," + "\n" 
                   ":cust_id," + "\n" 
                   ":cust_name);" + "\n"
                   )

    fl_file.write ("END LOADING;" + "\n")
    fl_file.write ("\n")
    fl_file.write ("LOG OFF;" + "\n")
    print ("FASTLOAD SCRIPT COMPLETE. ")


'''
for key , value in rca_data.items():
    print ("Key is " + str(key) + " --> Value is " + str(value).lstrip('[').rstrip(']'))

'''


if __name__ == "__main__":
    create_data_file(i)
    create_fl_script()


